import openpyxl
from io import BytesIO
from app.schemas import ExportRequest

def build_xlsx(req: ExportRequest) -> bytes:
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # 1. 00_Decision
    ws_dec = wb.create_sheet("00_Decision")
    dec_cols = ["job_id", "verdict", "approved_by", "approved_at", "rule_version", "parser_version", "final_recon_status", "zero_count", "amber_count", "prism_kernel_proof_ref", "costguard_band_summary", "watermark", "generated_at"]
    ws_dec.append(dec_cols)
    sorted_dec = sorted(req.decision_rows, key=lambda x: x.job_id or "")
    for row in sorted_dec:
        ws_dec.append([
            row.job_id,
            row.verdict,
            row.approved_by,
            row.approved_at,
            row.rule_version,
            row.parser_version,
            row.final_recon_status,
            row.zero_count,
            row.amber_count,
            row.prism_kernel_proof_ref,
            row.costguard_band_summary,
            row.watermark,
            row.generated_at or req.generated_at
        ])

    # 2. 01_Action_Items
    ws_act = wb.create_sheet("01_Action_Items")
    act_cols = ["action_id", "severity", "shipment_ref", "line_id", "issue_type", "required_action", "owner_role", "status", "prism_kernel_proof_ref"]
    ws_act.append(act_cols)
    sorted_act = sorted(req.action_items_rows, key=lambda x: x.action_id or "")
    for row in sorted_act:
        ws_act.append([
            row.action_id,
            row.severity,
            row.shipment_ref,
            row.line_id,
            row.issue_type,
            row.required_action,
            row.owner_role,
            row.status,
            row.prism_kernel_proof_ref
        ])

    # 3. 02_Final_Recon
    ws_recon = wb.create_sheet("02_Final_Recon")
    recon_cols = ["currency", "shipment_ref", "invoice_total", "reviewed_total", "variance", "variance_pct", "recon_status", "evidence_ref"]
    ws_recon.append(recon_cols)
    sorted_recon = sorted(req.final_recon_rows, key=lambda x: (x.shipment_ref or "", x.currency or ""))
    for row in sorted_recon:
        ws_recon.append([
            row.currency,
            row.shipment_ref,
            row.invoice_total,
            row.reviewed_total,
            row.variance,
            row.variance_pct,
            row.recon_status,
            row.evidence_ref
        ])

    # 3. 03_Type_B_Summary (ported from v3.2 SHPT full package - mandatory per Release_Gate and 8-sheet contract)
    # Columns from package: Shipment_No, Customs, DO, INLAND, THC, Inspection, Detention, STROAGE, OTHERS, Total_AED, Total_USD, Line_Count, Evidence_Status_Summary, Risk, Evidence
    ws_typeb = wb.create_sheet("03_Type_B_Summary")
    typeb_cols = ["Shipment_No", "Customs", "DO", "INLAND", "THC", "Inspection", "Detention", "STROAGE", "OTHERS", "Total_AED", "Total_USD", "Line_Count", "Evidence_Status_Summary", "Risk", "Evidence"]
    ws_typeb.append(typeb_cols)
    # Aggregate from line_view_rows by shipment and type_b (TYPE_B buckets from rules: Customs, DO, INLAND, THC, Inspection, Detention, STROAGE, OTHERS)
    from collections import defaultdict
    typeb_agg = defaultdict(lambda: defaultdict(float))
    typeb_lines = defaultdict(int)
    typeb_ev = defaultdict(set)
    for row in req.line_view_rows:
        ship = row.shipment_ref or "UNKNOWN"
        tb = (row.type_b or "OTHERS").upper()
        if tb not in ["CUSTOMS", "DO", "INLAND", "THC", "INSPECTION", "DETENTION", "STROAGE", "OTHERS"]:
            tb = "OTHERS"
        amt = row.amount or 0
        typeb_agg[ship][tb] += amt
        typeb_lines[ship] += 1
        if row.evidence_status:
            typeb_ev[ship].add(row.evidence_status)
    for ship in sorted(typeb_agg.keys()):
        d = typeb_agg[ship]
        total_aed = sum(d.values())
        total_usd = total_aed / 3.6725  # approx FX as in SHPT rules
        ev_sum = ",".join(sorted(typeb_ev[ship])) if typeb_ev[ship] else ""
        ws_typeb.append([
            ship,
            d.get("CUSTOMS", 0),
            d.get("DO", 0),
            d.get("INLAND", 0),
            d.get("THC", 0),
            d.get("INSPECTION", 0),
            d.get("DETENTION", 0),
            d.get("STROAGE", 0),
            d.get("OTHERS", 0),
            total_aed,
            round(total_usd, 2),
            typeb_lines[ship],
            ev_sum,
            "",  # Risk placeholder
            ""   # Evidence placeholder
        ])

    # 4. 04_Line_View
    ws_line = wb.create_sheet("04_Line_View")
    line_cols = ["line_id", "shipment_ref", "description", "for_charge_component", "type_b", "amount", "currency", "rate_source", "rate_status", "validity_status", "evidence_status", "gate_status", "band", "delta_pct", "numeric_integrity_status", "difference"]
    ws_line.append(line_cols)
    sorted_lines = sorted(req.line_view_rows, key=lambda x: x.line_id or "")
    for row in sorted_lines:
        ws_line.append([
            row.line_id,
            row.shipment_ref,
            row.description,
            row.for_charge_component,
            row.type_b,
            row.amount,
            row.currency,
            row.rate_source,
            row.rate_status,
            row.validity_status,
            row.evidence_status,
            row.gate_status,
            row.band,
            row.delta_pct,
            row.numeric_integrity_status,
            row.difference
        ])

    # 5. 90_Source_Data
    ws_src = wb.create_sheet("90_Source_Data")
    src_cols = ["file_id", "source_ref", "original_text", "normalized_value", "confidence", "routing_pattern", "pdf_page", "text_span_hash"]  # P3C
    ws_src.append(src_cols)
    sorted_src = sorted(req.source_data_rows, key=lambda x: (x.file_id or "", x.source_ref or ""))
    for row in sorted_src:
        ws_src.append([
            row.file_id,
            row.source_ref,
            row.original_text,
            row.normalized_value,
            row.confidence,
            row.routing_pattern,
            getattr(row, 'pdf_page', None),
            getattr(row, 'text_span_hash', None)
        ])

    # 6. 91_Audit_Detail
    ws_audit = wb.create_sheet("91_Audit_Detail")
    audit_cols = ["line_id", "rule_id", "reason_code", "sct_trace_id", "cf_mcp_tool", "cf_mcp_latency_ms", "confidence", "decision_input", "fx_override", "fx_policy_id"]
    ws_audit.append(audit_cols)
    sorted_audit = sorted(req.audit_detail_rows, key=lambda x: (x.line_id or "", x.rule_id or ""))
    for row in sorted_audit:
        ws_audit.append([
            row.line_id,
            row.rule_id,
            row.reason_code,
            row.sct_trace_id,
            row.cf_mcp_tool,
            row.cf_mcp_latency_ms,
            row.confidence,
            row.decision_input,
            row.fx_override,
            row.fx_policy_id
        ])

    # 7. 92_Evidence_Issues
    ws_ev = wb.create_sheet("92_Evidence_Issues")
    ev_cols = ["line_id", "required_evidence", "matched_evidence", "gap_type", "severity", "action_item_id", "human_gate_trigger_id"]
    ws_ev.append(ev_cols)
    sorted_ev = sorted(req.evidence_issues_rows, key=lambda x: (x.line_id or "", x.gap_type or ""))
    for row in sorted_ev:
        ws_ev.append([
            row.line_id,
            row.required_evidence,
            row.matched_evidence,
            row.gap_type,
            row.severity,
            row.action_item_id,
            row.human_gate_trigger_id
        ])

    # v3.2 SHPT port: add freeze + autofilter for Release_Gate / workbook_output_validate compliance (clears 'freeze top row not detected' / 'autofilter not detected' warnings)
    for ws in wb.worksheets:
        try:
            ws.freeze_panes = 'A2'
            if getattr(ws, 'max_row', 0) >= 1 and getattr(ws, 'max_column', 0) >= 1:
                ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

    out = BytesIO()
    wb.save(out)
    # v3.2 SHPT Release_Gate note (from DSV_SHIPMENT_FULL_PACKAGE_v3_2_PRO_INTERNAL): This now produces the exact 8-sheet contract per Release_Gate_v3.2 and 01_GPT_INSTRUCTIONS (00_Decision, 01_Action_Items, 02_Final_Recon, 03_Type_B_Summary, 04_Line_View, 90_Source_Data, 91_Audit_Detail, 92_Evidence_Issues).
    # Additional from v3.2: harness scripts (harness_validate_package.py, workbook_output_validate.py, run_self_test_3x.py) ported as quality gates in SCT/scripts/shpt_* for output contract enforcement, DLP, RTM, golden cases.
    # Full source_data from SHPT PDF spans + doc mapping (from port script and pdf_text SHPT extract) ensures reviewer feedback on complete population.
    # See port_shpt_fullset_gaps.py for v3.2 reference and SHPT 특화 (PortalFee, Gates, mapping).
    return out.getvalue()
